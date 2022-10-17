"""
MIDP Check

This script allows automated check of drawings title block info 
(Drawings numbers,Drawings titles, Drawing status, Suitability
and Revision) in an MIDP with pdf copies of the drawings.

The code produces a new excel file "MIDP_check_list.xlsx" with the check results.
Discrepancy in a drawing title is highlighted in YELLOW in this file and a
missing drawing number in MIDP is highlighted in RED.

For this script to work, it is required that WORK IN PROGRES RED BANNER is removed
from drawings before running this script.

It is recommended to save the drawings to be checked on a local drive (i.e. not
in Teams/SharePoint folder)

"""
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pdfminer.layout import LAParams, LTTextBox
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from tkinter import filedialog, messagebox
import tkinter as tk
import pdfplumber
import PyPDF2
import os
import gc
import warnings
warnings.filterwarnings("ignore")


def time_now():
    time_now = datetime.datetime.now()
    time = time_now - datetime.timedelta(microseconds=time_now.microsecond)
    return time


def list_dir(folder_path):
    for name in os.listdir(folder_path):
        path = os.path.join(folder_path, name)
        if os.path.isdir(path):
            yield from list_dir(path)
        else:
            file_path = (os.path.join(folder_path, name))
            yield file_path


def page_rotate(file, angle=90):
    with open(file, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfFileReader(pdf_file, strict=False)
        pdf_writer = PyPDF2.PdfFileWriter()
        for page_num in range(pdf_reader.numPages):
            pdf_page = pdf_reader.getPage(page_num)
            OrientationDegrees = pdf_page.get('/Rotate')
            print('default orientation -', OrientationDegrees)
            if OrientationDegrees not in [None, 0]:
                angle = -1*OrientationDegrees
                pdf_page.rotateClockwise(angle)
            dims = pdf_page.mediaBox
            page_height = dims.getHeight()
            page_width = dims.getWidth()
            if page_height > page_width:
                pdf_page.rotateClockwise(angle)
            pdf_writer.addPage(pdf_page)
            with open("rotated.pdf", 'wb') as pdf_file_rotated:
                pdf_writer.write(pdf_file_rotated)
    del pdf_file
    gc.collect()
    os.remove(file)
    os.rename("rotated.pdf", file)


def get_coordinates(file):
    fp = open(file, 'rb')
    rsrcmgr = PDFResourceManager()
    laparams = LAParams()
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    pages = PDFPage.get_pages(fp)
    word_list = ['Drawing title', 'Drawing  Number']
    coordinates_dict = {}
    for page in pages:
        interpreter.process_page(page)
        layout = device.get_result()
        for lobj in layout:
            if isinstance(lobj, LTTextBox):
                x, y, text = lobj.bbox[0], lobj.bbox[3], lobj.get_text()
                if text.strip() in word_list:
                    coordinates_dict[text.strip()] = (x, y)
    return coordinates_dict


def drawing_titleblock(file):
    Titleblock = []
    Word_coordinates = get_coordinates(file)
    x1 = Word_coordinates['Drawing title'][0]
    y1 = Word_coordinates['Drawing  Number'][1]
    y2 = Word_coordinates['Drawing title'][1]
    with pdfplumber.open(file) as pdf:
        pages = pdf.pages
        for i, j in enumerate(pages):
            page = pdf.pages[i]
            crop = page.crop((x1, page.height-y2, page.width, page.height-(y1-20)))
            data = crop.extract_text()
            value_0 = data.find('Drawing title')
            value_1 = data.find('Drawing Status')
            value_2 = data.find('Status Code')
            value_3 = data.find('Scale')
            value_4 = data.find('DO NOT SCALE')
            value_5 = data.find('Jacobs No.')
            value_6 = data.find('Client No.')
            value_6a = data.find('Revision Code')
            value_7 = data.find('Drawing  Number')
            value_7a = data.find('Model Reference')
            Drawing_title = data[(value_0+13):(value_1)].strip()
            Drawing_title = Drawing_title.replace("\n", " ")
            Drawing_status = data[(value_1+14):(value_2)].strip()
            Status_code = data[(value_2+12):(value_3)].strip()
            Scale = data[(value_3+5):(value_5)].strip()
            Jacobs_No = data[(value_5+10):(value_4)].strip()
            Client_No = data[(value_6+10):(value_6a)].strip()
            Drawing_number = data[(value_7+16):(value_7+46)].strip()
            Rev = data[(value_6a+14):(value_7a)].strip()
            Title_Blockinfo = [Drawing_number, Drawing_title, Drawing_status, Status_code, Scale, Jacobs_No, Client_No, Rev]
            for index, item in enumerate(Title_Blockinfo):
                Titleblock.append(item)
    return Titleblock


def column_width(column_dict, column_index, element):
    if column_dict[column_index] >= len(element):
        pass
    else:
        column_dict[column_index] = len(element)
    return column_dict


def to_excel(Titleblocks_list, DrawingsList_file):
    print('Writing excel file')
    workbook = Workbook()
    sheet = workbook.active
    column_dict = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0}
    for row, data in enumerate(Titleblocks_list):
        for column, element in enumerate(data):
            column_dict = column_width(column_dict, column, element)
            sheet.cell(row=row+1, column=column+1).value = element
            sheet.column_dimensions[get_column_letter(column+1)].width = column_dict[column]+2
            sheet.cell(row=row+1, column=column+1).alignment = Alignment(horizontal="center")
    print('Saving excel file')
    workbook.save(filename=DrawingsList_file)


def legends(excel_sheet):
    excel_sheet.cell(row=1, column=10).value = 'Legends'
    excel_sheet.cell(row=2, column=10).fill = PatternFill("solid", fgColor="00FFFF00")
    excel_sheet.cell(row=2, column=11).value = 'Value does not match with MIDP'
    excel_sheet.cell(row=3, column=10).fill = PatternFill("solid", fgColor="00FF0000")
    excel_sheet.cell(row=3, column=11).value = 'Drawing Number is not present in MIDP'
    excel_sheet.cell(row=4, column=11).value = 'Text style does not comply with Jacobs\' standards'
    excel_sheet.cell(row=4, column=11).font = Font(color='00FF0000')

def MIDP_info(MIDP_file, sheet_title, drawing_no, status=0):
    data = pd.ExcelFile(MIDP_file)
    sheet_title = sheet_title
    sheet_1 = pd.read_excel(data, sheet_title)
    column_names = sheet_1.iloc[2, :]
    sheet_1.columns = column_names
    if sheet_1.isin([drawing_no]).any().any():
        status = 1
    drawing_info = sheet_1[['Document Title', 'Status', 'RevCode']].where(sheet_1['Document Reference'] == drawing_no)
    drawing_info = drawing_info.dropna()
    drawing_info = drawing_info.values
    return drawing_info, status


def MIDP_check(DrawingsList_file, MIDP_file):
    print('Checking MIDP', time_now())
    Drawings_list = load_workbook(DrawingsList_file)
    DL_sheet = Drawings_list.active
    MIDP = load_workbook(MIDP_file, read_only=True, data_only=True)
    dict_values = []
    excluded_sheets = ['Project Environment', 'Format', 'Information Management', 'General Project Data', 'Spare-DoNotDelete', '_Lookup', 'Category_and_Number']
    for sheet in MIDP.worksheets:
        if sheet.title not in excluded_sheets:
            sheet_title = sheet.title
    for row_no in range(2, DL_sheet.max_row+1):
        MIDP_drawing_number = 'missing'
        Drawing_no = DL_sheet.cell(row=row_no, column=1).value
        DL_sheet_title = ' '.join(DL_sheet.cell(row=row_no, column=2).value.split())
        if any(letter.islower() for letter in DL_sheet_title):
            DL_sheet.cell(row=row_no, column=2).font = Font(color='00FF0000')
        for sheet in MIDP.worksheets:
            if sheet.title not in excluded_sheets:
                sheet_title = sheet.title
                drawing_info = MIDP_info(MIDP_file, sheet_title, Drawing_no)
                try:
                    if drawing_info[0][0][0].strip().lower() != DL_sheet_title.strip().lower():
                        DL_sheet.cell(row=row_no, column=2).fill = PatternFill("solid", fgColor="00FFFF00")
                    if DL_sheet.cell(row=row_no, column=3).value.strip().lower() not in drawing_info[0][0][1].lower():
                        DL_sheet.cell(row=row_no, column=3).fill = PatternFill("solid", fgColor="00FFFF00")
                    if DL_sheet.cell(row=row_no, column=4).value.strip().lower() not in drawing_info[0][0][1].split('-')[0].lower():
                        DL_sheet.cell(row=row_no, column=4).fill = PatternFill("solid", fgColor="00FFFF00")
                    if drawing_info[0][0][2].lower() != DL_sheet.cell(row=row_no, column=8).value.strip().lower():
                        DL_sheet.cell(row=row_no, column=8).fill = PatternFill("solid", fgColor="00FFFF00")
                    if drawing_info[1] == 1:
                        MIDP_drawing_number = 'present'
                except Exception as e:
                    # print(e)
                    pass
        if MIDP_drawing_number == 'missing':
            DL_sheet.cell(row=row_no, column=1).fill = PatternFill("solid", fgColor="00FF0000")

    legends(DL_sheet)
    Drawings_list.save(DrawingsList_file)
    print('Done', time_now())


def main():
    print('\n*** Select drawings folder ***\n')
    root = tk.Tk()
    root.attributes('-topmost', 1)
    root.update()
    root.withdraw()
    messagebox.showinfo(title='Drawings Check', message='Please ensure WORK IN PROGRES RED BANNER is removed from drawings before progressing')
    folder_path = filedialog.askdirectory(title='Select drawings folder')
    print('*** Select MIDP file ***\n')
    MIDP_file = filedialog.askopenfilename(filetypes=[('Microsoft Excel Worksheet', '*.xlsm')], title='Select MIDP file')
    t1 = time_now()
    print('Start -', t1)
    DrawingsList_file = os.path.join(folder_path, 'MIDP_check_list.xlsx')
    Titleblocks_list = [['Drawing Number', 'Drawing Title', 'Drawing Status', 'Suitability', 'Scale', 'Jacobs No', 'Client No', 'Rev']]
    file_path = list_dir(folder_path)
    while True:
        try:
            file = next(file_path)
            file_extension = (os.path.splitext(file)[1])
            if file_extension == '.pdf':
                print(file)
                try:
                    Titleblock = drawing_titleblock(file)
                    Titleblocks_list.append(Titleblock)
                except ValueError:
                    try:
                        page_rotate(file, 90)
                        print('Page rotated')
                        Titleblock = drawing_titleblock(file)
                        Titleblocks_list.append(Titleblock)
                    except ValueError:
                        page_rotate(file, -90)
                        print('Page rotated again')
                        Titleblock = drawing_titleblock(file)
                        Titleblocks_list.append(Titleblock)
        except StopIteration: break
    to_excel(Titleblocks_list, DrawingsList_file)
    MIDP_check(DrawingsList_file, MIDP_file)
    message = 'MIDP Drawings check list file is saved in this folder\n\n' + DrawingsList_file
    print('\n', message)
    print('Total time -', time_now()-t1)
    messagebox.showinfo(title='MIDP Check Complete', message=message)
    root.destroy()


if __name__ == "__main__": main()
